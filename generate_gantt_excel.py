import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = openpyxl.Workbook()
ws = wb.active
ws.title = "Gantt Chart"

tasks = [
    ("Architecture",     "Critical Path",   1,    1.0),
    ("Backend Core",     "Critical Path",   2,    1.0),
    ("Azure API",        "Critical Path",   3,    2.0),
    ("Data Sync",        "Critical Path",   5,    1.5),
    ("Anomaly Detect",   "Critical Path",   6.5,  1.0),
    ("UI Integration",   "Critical Path",   7.5,  1.5),
    ("System QA",        "Critical Path",   9,    1.5),
    ("Azure Infra",      "Infrastructure",  2,    1.0),
    ("CI/CD Pipeline",   "Infrastructure",  3,    1.0),
    ("Database",         "Data & Auth",     2,    1.0),
    ("Auth Module",      "Data & Auth",     3,    1.0),
    ("Frontend Core",    "Frontend / UI",   2,    0.5),
    ("Dashboard UI",     "Frontend / UI",   2.5,  2.0),
    ("Resource Mgmt",    "Frontend / UI",   4.5,  1.5),
    ("Budget Alerts",    "Analytics",       6.5,  1.0),
    ("Rec Engine",       "Analytics",       6.5,  1.0),
]

STEP = 0.5
TOTAL_WEEKS = 11
COLS = int(TOTAL_WEEKS / STEP)

COL_TASK  = 1
COL_TRACK = 2
COL_BAR   = 3

# ── Styles — NO fills, pure black & white ─────────────────────────────────────
thin   = Side(style="thin",   color="000000")
bdr    = Border(left=thin, right=thin, top=thin, bottom=thin)
no_fill = PatternFill(fill_type=None)   # completely no fill

font_bold  = Font(bold=True,  color="000000", size=12)
font_hdr   = Font(bold=True,  color="000000", size=11)
font_bar   = Font(bold=True,  color="000000", size=11)  # █ character
font_plain = Font(color="000000", size=11)
center     = Alignment(horizontal="center", vertical="center")
left       = Alignment(horizontal="left",   vertical="center")

# ── Header row ────────────────────────────────────────────────────────────────
ws.row_dimensions[1].height = 28

for ci, label in enumerate(["Task Name", "Track"], start=1):
    c = ws.cell(row=1, column=ci, value=label)
    c.fill      = no_fill
    c.font      = font_hdr
    c.alignment = center
    c.border    = bdr

for col_i in range(COLS):
    w = col_i * STEP + STEP
    label = f"W{int(w)}" if w == int(w) else f"W{w}"
    c = ws.cell(row=1, column=COL_BAR + col_i, value=label)
    c.fill      = no_fill
    c.font      = font_hdr
    c.alignment = center
    c.border    = bdr
    ws.column_dimensions[get_column_letter(COL_BAR + col_i)].width = 4

ws.column_dimensions[get_column_letter(COL_TASK)].width  = 24
ws.column_dimensions[get_column_letter(COL_TRACK)].width = 18

# ── Task rows ─────────────────────────────────────────────────────────────────
for row_i, (name, track, start, dur) in enumerate(tasks):
    row = row_i + 2
    ws.row_dimensions[row].height = 22

    # Task name
    c = ws.cell(row=row, column=COL_TASK, value=name)
    c.fill = no_fill; c.font = font_bold
    c.alignment = left; c.border = bdr

    # Track
    c = ws.cell(row=row, column=COL_TRACK, value=track)
    c.fill = no_fill; c.font = font_plain
    c.alignment = center; c.border = bdr

    # Week bar cells — █ for active, empty for inactive
    for col_i in range(COLS):
        week_val = round(col_i * STEP + STEP, 1)
        in_bar   = (start <= week_val < round(start + dur, 1))
        c = ws.cell(row=row, column=COL_BAR + col_i,
                    value="█" if in_bar else "")
        c.fill      = no_fill
        c.font      = font_bar
        c.alignment = center
        c.border    = bdr

ws.freeze_panes = "C2"

out = r'C:\Users\Lenovo\OneDrive\Desktop\Collage wla\GANTT-CHART.xlsx'
wb.save(out)
print(f"Saved: {out}")
